import os
from openpyxl.workbook import Workbook
from openpyxl.styles import Border, Side, Alignment, Font


# This is to ask for the directory path at the command prompt
dir_path = input('What is the full directory path? ')


def product_list_with_notes():
    """
    Generates a file listing all of the products in the job that have special notes
    Opens each of the .des files inside a job directory and looks at every product in the room
    If the product has notes, it adds that product to the list
    Then it writes the list of products with notes to a new file
    """
    dir_path_lst = dir_path.split('\\')
    job_name = dir_path_lst[-1]

    xml_char_ents = [
        ['&quot;', '"'],
        ['&#xD;&#xA;', '\n'],
        ['&amp;', '&'],
        ['&apos;', "'"],
        ['&gt;', '>'],
        ['&lt;', '<']
    ]

    prod_dict = {}
    
    for root, dirs, files in os.walk(dir_path):
        
        for file in files:

            if file.endswith('.des'):

                full_path = dir_path + '\\' + file
                f = open(full_path, "rt")
                content = f.readlines()
                rm_start_idx = content[2].find('Name=') + 6
                rm_end_idx = content[2].find('" RoomNosDirty=')
                room_name = content[2][rm_start_idx:rm_end_idx]
                room_num = int(file[4:file.find('.')])
                prod_dict[room_num] = [room_name,[]]
                                
                for line in content:
                    if line.startswith('    <Product '):
                        note_start_idx = line.find('Notes=') + 6
                        note_end_idx = line.find('" Price=')
                        note_intro = line[note_start_idx:note_start_idx + 2]

                        if note_intro != '""':
                            prod_start_idx = line.find('ProdName=') + 10
                            prod_end_idx = line.find('" IDTag=')
                            prod_name = line[prod_start_idx:prod_end_idx]
                            note = line[note_start_idx + 1:note_end_idx]

                            for char in xml_char_ents:
                                note = note.replace(char[0], char[1])
                            
                            prod_dict[room_num][1].append([prod_name, note])





    # print statement to see clean list of lists (requires import json)
    # print(json.dumps(prod_dict, indent=4))

    wb = Workbook()
    sheet1 = wb.active
    row = 1
    col = 1
    room_key = 0

    sheet1.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col + 1)
    sheet1.cell(row, col, job_name + ' - Special Notes')
    sheet1.cell(row, col).alignment = Alignment(vertical='top')
    sheet1.cell(row, col).font = Font(size=12, bold=True, italic=True)
    sheet1.row_dimensions[1].height = 25
    row += 1

    for i in range(len(prod_dict)):
        if prod_dict[room_key][1] != []:
            sheet1.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col + 1)
            sheet1.cell(row, col, prod_dict[room_key][0])
            sheet1.cell(row, col).alignment = Alignment(wrapText=True, horizontal='center')
            sheet1.cell(row, col).font = Font(size=14, bold=True, underline='single')
            row += 1

            sheet1.cell(row, col, 'Product Name')
            sheet1.cell(row, col).alignment = Alignment(horizontal='general', indent=2.0)
            sheet1.cell(row, col).font = Font(size=12, italic=True)
            sheet1.cell(row, col).border = Border(bottom=Side(style='thin', color='000000'))
            col += 1

            sheet1.cell(row, col, 'Notes')
            sheet1.cell(row, col).alignment = Alignment(horizontal='general', indent=2.0)
            sheet1.cell(row, col).font = Font(size=12, italic=True)
            sheet1.cell(row, col).border = Border(bottom=Side(style='thin', color='000000'))
            col -= 1
            row += 1


            for prod in prod_dict[room_key][1]:
                sheet1.cell(row, col, prod[0])
                sheet1.cell(row, col).alignment = Alignment(wrapText=True, vertical='top', indent=2.0)
                sheet1.cell(row, col).font = Font(size=11)
                sheet1.cell(row, col).border = Border(bottom=Side(style='thin', color='D4D4D4'))
                col += 1
                sheet1.cell(row, col, prod[1])
                sheet1.cell(row, col).alignment = Alignment(wrapText=True, vertical='top', indent=2.0)
                sheet1.cell(row, col).font = Font(size=11)
                sheet1.cell(row, col).border = Border(bottom=Side(style='thin', color='D4D4D4'))
                col -= 1
                row += 1

            row += 1
        room_key += 1

    sheet1.column_dimensions['A'].width = 25
    sheet1.column_dimensions['B'].width = 58
    
    print_area = 'A1:B' + str(row)
    sheet1.print_area = print_area
    
    save_name = job_name + '.xlsx'
    full_save_name = os.path.join(dir_path, save_name)
    try:
        wb.save(full_save_name)
    except PermissionError:
        print("\nSAVE FAILED\nYou will need to close the open file before it can be saved.")

    os.startfile(full_save_name)
    
product_list_with_notes()
