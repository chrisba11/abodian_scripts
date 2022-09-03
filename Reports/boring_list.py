import os
import pprint
from openpyxl.workbook import Workbook
from openpyxl.styles import Border, Side, Alignment, Font
from openpyxl.worksheet.pagebreak import Break
from datetime import datetime


# This is to ask for the directory path at the command prompt
dir_path = os.path.normpath(input('What is the full directory path for the job? '))
rooms_string = input('Which rooms should be included in this report? List the room integers only, separated by commas and no spaces. Include "0" if you want Order Entry included. If all rooms, type "all". ').lower()
rooms_in_report = rooms_string.split(',')

now = datetime.now()
now_string = now.strftime("%m.%d.%Y-%H.%M.%S")

def mm_to_in(metric_decimal):
    inch_decimal = metric_decimal / 25.4
    inch_int = int(inch_decimal)
    inch_remain = inch_decimal % 1

    if inch_remain < 0.03125:
        inch_fraction = 0
    elif inch_remain < 0.09375:
        inch_fraction = 0.0625
    elif inch_remain < 0.15625:
        inch_fraction = 0.125
    elif inch_remain < 0.21875:
        inch_fraction = 0.1875
    elif inch_remain < 0.28125:
        inch_fraction = 0.25
    elif inch_remain < 0.34375:
        inch_fraction = 0.3125
    elif inch_remain < 0.40625:
        inch_fraction = 0.375
    elif inch_remain < 0.46875:
        inch_fraction = 0.4375
    elif inch_remain < 0.53125:
        inch_fraction = 0.5
    elif inch_remain < 0.59375:
        inch_fraction = 0.5625
    elif inch_remain < 0.65625:
        inch_fraction = 0.625
    elif inch_remain < 0.71875:
        inch_fraction = 0.6875
    elif inch_remain < 0.78125:
        inch_fraction = 0.75
    elif inch_remain < 0.84375:
        inch_fraction = 0.8125
    elif inch_remain < 0.90625:
        inch_fraction = 0.875
    elif inch_remain < 0.96875:
        inch_fraction = 0.9375
    else:
        inch_fraction = 0
        inch_int += 1

    inches = inch_int + inch_fraction

    return inches


class Room:
    def __init__(self, number, mat_door_template):
        self.number            = number
        self.mat_door_template = mat_door_template
        self.products          = []


class Product:
    def __init__(self, unique_id, name, cab_num, door_mat_override):
        self.unique_id         = unique_id
        self.name              = name
        self.cab_num           = cab_num
        self.door_mat_override = door_mat_override
        self.doors             = []
    
    def add_door(self, door):
        self.doors.append(door)
    
    def get_doors(self):
        return self.doors


class Door:
    def __init__(self, name=None, report_name=None, door_style=None,
                width=None, height=None, comment=None, quantity=0,
                hinge_type=None, hinge_edge=None,
                hinge_centers=None, horizontal_grain=False):

        self.name          = name
        self.report_name   = report_name
        self.comment       = comment
        self.door_style    = door_style
        self.width         = width
        self.height        = height
        self.quantity      = quantity
        self.hinge_type    = hinge_type
        self.hinge_edge    = hinge_edge
        self.hinge_centers = hinge_centers
        self.horizontal    = horizontal_grain


class Material:
    def __init__(self, name):
        self.name = name


class DoorStyle:
    def __init__(self, name):
        self.name = name


class HingeType:
    def __init__(self, name):
        self.name = name


def hinge_boring_report():
    """

    """
    dir_path_lst = dir_path.split('\\')
    job_name = dir_path_lst[-1]

    product_dict = {}
    is_blind = False

    room_list = []


# #    product_dict sample
#     product_dict = {
#         "Room1": {
#             "MatDoorTemplate": "02 - 5pc Painted - In House MDF (1)",
#             "Products": {
#                 "C4": {
#                     "UniqueID": "3101352",
#                     "Name": "Base Sink",
#                     "MatOR": None,
#                     "Doors": [
#                         {
#                             "DoorStyle": "Shaker Door",
#                             "Name": "Door(L)",
#                             "ReportName": "Door(L)",
#                             "Comment": "Cabinet Door",
#                             "Quan": 1,
#                             "W": 454.0125,
#                             "H": 771.5,
#                             "HingeCenterLines": [101.6, 542.9],
#                             "HingeEdge": "Left",
#                             "HingeType": "DTC C-80 110 / 2mm - 105-C80A675NF",
#                             "IsHorizGrain": False,
#                         },
#                         {
#                             "DoorStyle": "Shaker Door",
#                             "Name": "Door(R)",
#                             "ReportName": "Door(R)",
#                             "Comment": "Cabinet Door",
#                             "Quan": 1,
#                             "W": 454.0125,
#                             "H": 771.5,
#                             "HingeCenterLines": [101.6, 542.9],
#                             "HingeEdge": "Right",
#                             "HingeType": "DTC C-80 110 / 2mm - 105-C80A675NF",
#                             "IsHorizGrain": False,
#                         },
#                     ]
#                 },
#                 "C6": {
#                     "UniqueID": "3101389",
#                     "Name": "Base Corner Left - Susan (L)",
#                     "MatOR": None,
#                     "Doors": [
#                         {
#                             "DoorStyle": "Shaker Door",
#                             "Name": "Door(L)",
#                             "ReportName": "Door(L)",
#                             "Comment": "Cabinet Door",
#                             "Quan": 1,
#                             "W": 454.0125,
#                             "H": 771.5,
#                             "HingeCenterLines": [101.6, 542.9],
#                             "HingeEdge": "Left",
#                             "HingeType": "DTC C-80 110 / 2mm - 105-C80A675NF",
#                             "Horizontal": False,
#                         },
#                         {
#                             "DoorStyle": "Shaker Door",
#                             "Name": "Door(L)",
#                             "ReportName": "Door(L)",
#                             "Comment": "Cabinet Door",
#                             "Quan": 1,
#                             "W": 454.0125,
#                             "H": 771.5,
#                             "HingeCenterLines": [101.6, 542.9],
#                             "HingeEdge": "Right",
#                             "HingeType": "DTC C-80 110 / 2mm - 105-C80A675NF",
#                             "Horizontal": False,
#                         },
#                     ]
#                 },
#             },
#         },
#     }

    
    for root, dirs, files in os.walk(dir_path):
        
        for file in files:

            if file.endswith('.des'):

                full_path = os.path.join(dir_path, file)
                with open(full_path, "rt") as f:
                    content = f.readlines()

                # room settings
                room_num = file[4:file.find('.')]
                # only record rooms requested for this report
                if room_num not in rooms_in_report and 'all' not in rooms_in_report:
                    continue
                mat_temp_start_idx = content[4].find('MatDoorTemplate=') + 17
                mat_temp_end_idx = content[4].find('" MatDrawerTemplate=')
                room_mat_door_template = content[4][mat_temp_start_idx:mat_temp_end_idx]

                ## instantiate room
                room = Room(number=room_num, mat_door_template=room_mat_door_template)
                ## add room to room list
                room_list.append(room)

                # set products var to current room's products list
                # products = product_dict[room_num]["Products"]

                # declaring variables that are used in multiple scopes
                door = None

                for line in content:
                    if line.startswith('    <Product '):
                        # product number
                        start_idx = line.find('CabNo=') + 7
                        end_idx = line.find('" Numbered=')
                        prod_num = line[start_idx:end_idx]
                        
                        # product number prefix
                        start_idx = end_idx + 12
                        num_prefix = 'C' if line[start_idx:start_idx + 1] == 'T' else 'N'
                        
                        # unique id
                        start_idx = line.find('UniqueID=') + 10
                        end_idx = line.find('" OrderID=')
                        unique_id = line[start_idx:end_idx]

                        # product name
                        start_idx = line.find('ProdName=') + 10
                        end_idx = line.find('" IDTag=')
                        prod_name = line[start_idx:end_idx]

                        # door material override
                        start_idx = line.find('MatDoorOR=') + 11
                        end_idx = line.find('" MatDrawerOR=')
                        door_mat_or = line[start_idx:end_idx] if line[start_idx:end_idx] != '' else None

                        # instantiate product
                        product = Product(unique_id, prod_name, prod_num, door_mat_or)
                        # add product to room's product list
                        room.products.append(product)


                    if line.startswith('        <ProductDoor '):
                        # instantiate door
                        door = Door()

                        # door style
                        start_idx = line.find('DoorStyle=') + 11
                        end_idx = line.find('" W=')
                        door_style = line[start_idx:end_idx]
                        door.door_style = door_style

                        # width
                        start_idx = end_idx + 5
                        end_idx = line.find('" H=')
                        width = float(line[start_idx:end_idx])
                        door.width = width
                        
                        # height
                        start_idx = end_idx + 5
                        end_idx = line.find('" Oversize=')
                        height = float(line[start_idx:end_idx])
                        door.height = height

                        # hinge center lines
                        start_idx = line.find('HingeCenterLines=') + 18
                        end_idx = line.find('" HingeEdge=')
                        hinge_centers = line[start_idx:end_idx].split('#')
                        hinge_centers = [float(i) for i in hinge_centers if i != '']
                        door.hinge_centers = hinge_centers

                        # hinge edge
                        start_idx = end_idx + 13
                        end_idx = line.find('" HingeType=')
                        hinge_edge = line[start_idx:end_idx]
                        door.hinge_edge = hinge_edge

                        # hinge type
                        start_idx = end_idx + 13
                        end_idx = line.find('" IsDrawerFront=')
                        hinge_type = line[start_idx:end_idx]
                        door.hinge_type = hinge_type

                        # horizontal grain?
                        start_idx = line.find('IsHorizGrain=') + 14
                        end_idx = line.find('" DoorType=')
                        horizontal = line[start_idx:end_idx]
                        horizontal = True if horizontal == 'True' else False
                        door.horizontal = horizontal


                    if line.startswith('          <DoorProdPart '):
                        # door name
                        start_idx = line.find('Name=') + 6
                        end_idx = line.find('" ReportName=')
                        door_name = line[start_idx:end_idx]
                        door.name = door_name
                        
                        # report name
                        start_idx = end_idx + 14
                        end_idx = line.find('" UsageType=')
                        report_name = line[start_idx:end_idx]
                        door.report_name = report_name

                        # comment
                        start_idx = line.find('Comment=') + 9
                        end_idx = line.find('" CommentLocked=')
                        comment = line[start_idx:end_idx]
                        door.comment = comment

                        # quantity
                        start_idx = line.find('Quan=') + 6
                        end_idx = line.find('" W=')
                        quantity = line[start_idx:end_idx]
                        door.quantity = int(quantity)


                    # If there is a door, append it to the list
                    if line.startswith('        </ProductDoor>'):
                        if door.quantity > 0:
                            product.doors.append(door)


                    # Check for blind panels and add them
                    if line.startswith('        <CabProdPart Name="Panel" ReportName=" Blind Panel'):
                        door = Door()
                        is_blind = True

                        # door name
                        start_idx = line.find('Name=') + 6
                        end_idx   = line.find('" ReportName=')
                        door_name = line[start_idx:end_idx]
                        door.name = door_name
                        
                        # report name
                        start_idx        = end_idx + 14
                        end_idx          = line.find('" UsageType=')
                        report_name      = line[start_idx:end_idx]
                        door.report_name = report_name

                        # comment
                        start_idx    = line.find('Comment=') + 9
                        end_idx      = line.find('" CommentLocked=')
                        comment      = line[start_idx:end_idx]
                        door.comment = comment

                        # quantity
                        start_idx     = line.find('Quan=') + 6
                        end_idx       = line.find('" W=')
                        quantity      = line[start_idx:end_idx]
                        door.quantity = quantity

                        # width
                        start_idx  = end_idx + 5
                        end_idx    = line.find('" L=')
                        width      = float(line[start_idx:end_idx])
                        door.width = width
                        
                        # length
                        start_idx   = end_idx + 5
                        end_idx     = line.find('" Color=')
                        height      = float(line[start_idx:end_idx])
                        door.height = height

                        # append defaults for blind panels
                        door.hinge_centers = ''
                        door.hinge_edge    = ''
                        door.hinge_type    = ''
                        door.horizontal    = False
                    

                    # If is_blind is True and there is a door, append it to the list
                    if line.startswith('        </CabProdPart>'):
                        if is_blind:
                            if door.quantity > 0:
                                product.doors.append(door)
                                is_blind = False


    ####
    # Start back here with converting to OOP
    ####

    sorted_product_dict = {}
    materials = set()
    door_styles = set()
    hinge_types = set()
    std_hinge_dist = 101.6

    for _room in product_dict:
        __room = product_dict[_room]
        mat = __room["MatDoorTemplate"]
        # if material does not exist in sorted product list, add it
        if mat not in sorted_product_dict:
            sorted_product_dict[mat] = {}
            materials.add(mat)

        for _product in __room["Products"]:
            __product = __room["Products"][_product]
            cab_num = 'R' + _room + _product

            mat = __room["MatDoorTemplate"]
            # if product has material override
            if __product["MatOR"] != None:
                mat = __product["MatOR"]

                # if material does not exist in sorted product list, add it
                if mat not in sorted_product_dict:
                    sorted_product_dict[mat] = {}
                    materials.add(mat)
            
            for _door in __product["Doors"]:
                mat_dict = sorted_product_dict[mat]
                door_style = _door["DoorStyle"]
                hinge_type = _door["HingeType"]

                # if door style is not in material dict, add it
                if door_style not in mat_dict:
                    mat_dict[door_style] = {}
                    door_styles.add(door_style)

                # if hinge type is not in door style dict, add it
                if hinge_type not in mat_dict[door_style]:
                    mat_dict[door_style][hinge_type] = []
                    hinge_types.add(hinge_type)
                
                # simplify access to door list for this door's hinge type
                door_list = mat_dict[door_style][hinge_type]

                # assign true hinge centers from top/bottom of door for up to 4 hinges
                hinge_centers = _door["HingeCenterLines"]
                if hinge_centers != '-':
                    num_hinges = len(hinge_centers)
                    bot_hinge_center = round(hinge_centers[0],1) if num_hinges > 0 else 0
                    bot_mid_hinge_center = round(hinge_centers[1],1) if num_hinges > 2 else 0
                    top_mid_hinge_center = round(hinge_centers[2],1) if num_hinges > 3 else 0
                    if (_door["HingeEdge"] == 'Left' or _door["HingeEdge"] == 'Right') and _door["IsHorizGrain"] == False:
                        top_hinge_center = round(_door["H"] - hinge_centers[-1],1) if num_hinges > 1 else 0
                    else:
                        top_hinge_center = round(_door["W"] - hinge_centers[-1],1) if num_hinges > 1 else 0
                else:
                    bot_hinge_center = '-'
                    bot_mid_hinge_center = '-'
                    top_mid_hinge_center = '-'
                    top_hinge_center = '-'


                # is the hinging standard or not
                std_not = "S"
                if bot_hinge_center != std_hinge_dist or top_hinge_center != std_hinge_dist:
                    std_not = "N"
                
                if bot_hinge_center == '-':
                    std_not = "-"
                
                door_details = [
                    _door["Quan"],
                    _door["Name"],
                    _door["Comment"],
                    _door["W"],
                    _door["H"],
                    std_not,
                    _door["HingeEdge"],
                    bot_hinge_center,
                    bot_mid_hinge_center,
                    top_mid_hinge_center,
                    top_hinge_center,
                    [cab_num]
                ]

                # if door with matching details does not exist in the list, add it
                # if a matching door does exist, add the cabinet number to the list of cab numbers
                # and increment the corresponding quantity in the quantity list
                found_match = False
                for door in door_list:
                    if door_details[1:-1] == door[1:-1]:
                        door[0] += door_details[0]
                        for cab_no in door_details[-1]:
                            door[-1].append(cab_no)
                        found_match = True
                
                if found_match is False:
                    door_list.append(door_details)

    # remove any materials that don't have doors associated with them
    temp_mat_set = set()
    temp_styles_set = set()
    temp_types_set = set()

    for _material in materials:
        for _style in door_styles:
            if _style not in sorted_product_dict[_material]:
                continue

            for _type in hinge_types:
                if _type not in sorted_product_dict[_material][_style]:
                    continue

                list_of_doors = sorted_product_dict[_material][_style][_type]

                if len(list_of_doors) > 0:
                    temp_mat_set.add(_material)
                    temp_styles_set.add(_style)
                    temp_types_set.add(_type)
    
    # creates sorted lists from sets
    mat_list = sorted(temp_mat_set)
    styles_list = sorted(temp_styles_set)
    types_list = sorted(temp_types_set)

    def create_excel():

        wb = Workbook()
        sheet1 = wb.active
        row = 1
        col = 1
        
        for _material in mat_list:
            sheet1.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col + 12)
            sheet1.cell(row, col, _material)
            sheet1.cell(row, col).alignment = Alignment(wrapText=True, vertical='center', horizontal='left')
            sheet1.cell(row, col).font = Font(size=16, bold=True)
            sheet1.row_dimensions[row].height = 18
            row += 2

            mat = sorted_product_dict[_material]

            mat_total = 0

            for _style in styles_list:
                if _style in mat:
                    door_style = mat[_style]
                else:
                    continue

                sheet1.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col + 12)
                sheet1.cell(row, col, _style)
                sheet1.cell(row, col).alignment = Alignment(wrapText=True, horizontal='left')
                sheet1.cell(row, col).font = Font(size=14, bold=True, underline='single')
                sheet1.row_dimensions[row].height = 18
                row += 2
                col += 1

                style_total = 0

                for _type in types_list:
                    if _type in door_style:
                        hinge = door_style[_type]
                    else:
                        continue

                    if len(hinge) == 0:
                        continue
                    
                    # Change printed type to say 'No Hinges' if there is no type listed
                    __type = _type if _type != '' else 'No Hinges'

                    sheet1.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col + 11)
                    sheet1.cell(row, col, __type)
                    sheet1.cell(row, col).alignment = Alignment(wrapText=True, vertical='top', horizontal='left')
                    sheet1.cell(row, col).font = Font(size=12, bold=True, italic=True)
                    sheet1.cell(row, col).border = Border(bottom=Side(style='thin', color='000000'))
                    sheet1.cell(row, col + 1).border = Border(bottom=Side(style='thin', color='000000'))
                    sheet1.cell(row, col + 2).border = Border(bottom=Side(style='thin', color='000000'))
                    sheet1.cell(row, col + 3).border = Border(bottom=Side(style='thin', color='000000'))
                    sheet1.cell(row, col + 4).border = Border(bottom=Side(style='thin', color='000000'))
                    sheet1.cell(row, col + 5).border = Border(bottom=Side(style='thin', color='000000'))
                    sheet1.cell(row, col + 6).border = Border(bottom=Side(style='thin', color='000000'))
                    sheet1.cell(row, col + 7).border = Border(bottom=Side(style='thin', color='000000'))
                    sheet1.cell(row, col + 8).border = Border(bottom=Side(style='thin', color='000000'))
                    sheet1.cell(row, col + 9).border = Border(bottom=Side(style='thin', color='000000'))
                    sheet1.cell(row, col + 10).border = Border(bottom=Side(style='thin', color='000000'))
                    sheet1.cell(row, col + 11).border = Border(bottom=Side(style='thin', color='000000'))
                    row += 1
                    sheet1.cell(row, col, "Qty")
                    sheet1.cell(row, col).alignment = Alignment(wrapText=True, vertical='center', horizontal='center')
                    sheet1.cell(row, col).font = Font(size=10, bold=True, italic=True)
                    sheet1.cell(row, col).border = Border(bottom=Side(style='thin', color='000000'))
                    col += 1
                    sheet1.cell(row, col, "Door Name")
                    sheet1.cell(row, col).alignment = Alignment(wrapText=True, vertical='center', horizontal='center')
                    sheet1.cell(row, col).font = Font(size=10, bold=True, italic=True)
                    sheet1.cell(row, col).border = Border(bottom=Side(style='thin', color='000000'))
                    col += 1
                    sheet1.cell(row, col, "Comment")
                    sheet1.cell(row, col).alignment = Alignment(wrapText=True, vertical='center', horizontal='center')
                    sheet1.cell(row, col).font = Font(size=10, bold=True, italic=True)
                    sheet1.cell(row, col).border = Border(bottom=Side(style='thin', color='000000'))
                    col += 1
                    sheet1.cell(row, col, "W")
                    sheet1.cell(row, col).alignment = Alignment(wrapText=True, vertical='center', horizontal='center')
                    sheet1.cell(row, col).font = Font(size=10, bold=True, italic=True)
                    sheet1.cell(row, col).border = Border(bottom=Side(style='thin', color='000000'))
                    col += 1
                    sheet1.cell(row, col, "H")
                    sheet1.cell(row, col).alignment = Alignment(wrapText=True, vertical='center', horizontal='center')
                    sheet1.cell(row, col).font = Font(size=10, bold=True, italic=True)
                    sheet1.cell(row, col).border = Border(bottom=Side(style='thin', color='000000'))
                    col += 1
                    sheet1.cell(row, col, "Std")
                    sheet1.cell(row, col).alignment = Alignment(wrapText=True, vertical='center', horizontal='center')
                    sheet1.cell(row, col).font = Font(size=10, bold=True, italic=True)
                    sheet1.cell(row, col).border = Border(bottom=Side(style='thin', color='000000'))
                    col += 1
                    sheet1.cell(row, col, "Edge")
                    sheet1.cell(row, col).alignment = Alignment(wrapText=True, vertical='center', horizontal='center')
                    sheet1.cell(row, col).font = Font(size=10, bold=True, italic=True)
                    sheet1.cell(row, col).border = Border(bottom=Side(style='thin', color='000000'), right=Side(style='thin', color='D4D4D4'))
                    col += 1
                    sheet1.cell(row, col, "BOT\nor LT")
                    sheet1.cell(row, col).alignment = Alignment(wrapText=True, vertical='center', horizontal='center')
                    sheet1.cell(row, col).font = Font(size=10, bold=True, italic=True)
                    sheet1.cell(row, col).border = Border(bottom=Side(style='thin', color='000000'))
                    col += 1
                    sheet1.cell(row, col, "BOT2\nor LT2")
                    sheet1.cell(row, col).alignment = Alignment(wrapText=True, vertical='center', horizontal='center')
                    sheet1.cell(row, col).font = Font(size=10, bold=True, italic=True)
                    sheet1.cell(row, col).border = Border(bottom=Side(style='thin', color='000000'))
                    col += 1
                    sheet1.cell(row, col, "BOT3\nor LT3")
                    sheet1.cell(row, col).alignment = Alignment(wrapText=True, vertical='center', horizontal='center')
                    sheet1.cell(row, col).font = Font(size=10, bold=True, italic=True)
                    sheet1.cell(row, col).border = Border(bottom=Side(style='thin', color='000000'))
                    col += 1
                    sheet1.cell(row, col, "TOP\nor RT")
                    sheet1.cell(row, col).alignment = Alignment(wrapText=True, vertical='center', horizontal='center')
                    sheet1.cell(row, col).font = Font(size=10, bold=True, italic=True)
                    sheet1.cell(row, col).border = Border(bottom=Side(style='thin', color='000000'), right=Side(style='thin', color='D4D4D4'))
                    col += 1
                    sheet1.cell(row, col, "Cab #")
                    sheet1.cell(row, col).alignment = Alignment(wrapText=True, vertical='center', horizontal='center')
                    sheet1.cell(row, col).font = Font(size=10, bold=True, italic=True)
                    sheet1.cell(row, col).border = Border(bottom=Side(style='thin', color='000000'))
                    row += 1
                    col -= 11
                    type_total = 0

                    hinge.sort(key = lambda x: x[10])
                    hinge.sort(key = lambda x: x[7])
                    hinge.sort(key = lambda x: x[3])
                    hinge.sort(key = lambda x: x[4])
                    hinge.sort(key = lambda x: x[6])
                    hinge.sort(key = lambda x: x[5])
                    
                    for door in hinge:
                        # Quantity
                        sheet1.cell(row, col, door[0])
                        sheet1.cell(row, col).alignment = Alignment(wrapText=True, vertical='center', horizontal='center')
                        sheet1.cell(row, col).font = Font(size=9)
                        sheet1.cell(row, col).border = Border(bottom=Side(style='thin', color='D4D4D4'))
                        col += 1
                        # Name
                        sheet1.cell(row, col, door[1])
                        sheet1.cell(row, col).alignment = Alignment(wrapText=True, vertical='center', horizontal='center')
                        sheet1.cell(row, col).font = Font(size=9)
                        sheet1.cell(row, col).border = Border(bottom=Side(style='thin', color='D4D4D4'))
                        col += 1
                        # Comment
                        sheet1.cell(row, col, door[2])
                        sheet1.cell(row, col).alignment = Alignment(wrapText=True, vertical='center', horizontal='center')
                        sheet1.cell(row, col).font = Font(size=9)
                        sheet1.cell(row, col).border = Border(bottom=Side(style='thin', color='D4D4D4'))
                        col += 1
                        # Width
                        sheet1.cell(row, col, mm_to_in(door[3]))
                        sheet1.cell(row, col).number_format = '#  ##/##'
                        sheet1.cell(row, col).alignment = Alignment(wrapText=True, vertical='center', horizontal='left', indent=1)
                        sheet1.cell(row, col).font = Font(size=9)
                        sheet1.cell(row, col).border = Border(bottom=Side(style='thin', color='D4D4D4'))
                        col += 1
                        # Height
                        sheet1.cell(row, col, mm_to_in(door[4]))
                        sheet1.cell(row, col).number_format = '#  ##/##'
                        sheet1.cell(row, col).alignment = Alignment(wrapText=True, vertical='center', horizontal='left', indent=1)
                        sheet1.cell(row, col).font = Font(size=9)
                        sheet1.cell(row, col).border = Border(bottom=Side(style='thin', color='D4D4D4'))
                        col += 1
                        # Standard or Non-Standard
                        sheet1.cell(row, col, door[5])
                        sheet1.cell(row, col).alignment = Alignment(wrapText=True, vertical='center', horizontal='center')
                        sheet1.cell(row, col).font = Font(size=9)
                        sheet1.cell(row, col).border = Border(bottom=Side(style='thin', color='D4D4D4'))
                        col += 1
                        # Hinge Edge
                        sheet1.cell(row, col, door[6])
                        sheet1.cell(row, col).alignment = Alignment(wrapText=True, vertical='center', horizontal='center')
                        sheet1.cell(row, col).font = Font(size=9)
                        sheet1.cell(row, col).border = Border(bottom=Side(style='thin', color='D4D4D4'), right=Side(style='thin', color='D4D4D4'))

                        col += 1
                        # Bottom/Left Hinge
                        if door[7] == 0:
                            hinge_center = ''
                        elif door[7] == '-':
                            hinge_center = door[7]
                        else:
                            hinge_center = mm_to_in(door[7])
                        sheet1.cell(row, col).number_format = '#  ##/##'
                        sheet1.cell(row, col, hinge_center)
                        sheet1.cell(row, col).alignment = Alignment(wrapText=True, vertical='center', horizontal='center')
                        sheet1.cell(row, col).font = Font(size=9)
                        sheet1.cell(row, col).border = Border(bottom=Side(style='thin', color='D4D4D4'))
                        col += 1
                        # Bottom/Left Mid Hinge
                        if door[8] == 0:
                            hinge_center = ''
                        elif door[8] == '-':
                            hinge_center = door[8]
                        else:
                            hinge_center = mm_to_in(door[8])
                        sheet1.cell(row, col, hinge_center)
                        sheet1.cell(row, col).number_format = '#  ##/##'
                        sheet1.cell(row, col).alignment = Alignment(wrapText=True, vertical='center', horizontal='center')
                        sheet1.cell(row, col).font = Font(size=9)
                        sheet1.cell(row, col).border = Border(bottom=Side(style='thin', color='D4D4D4'))
                        col += 1
                        # Top/Right Mid Hinge
                        if door[9] == 0:
                            hinge_center = ''
                        elif door[9] == '-':
                            hinge_center = door[9]
                        else:
                            hinge_center = mm_to_in(door[9])
                        sheet1.cell(row, col, hinge_center)
                        sheet1.cell(row, col).number_format = '#  ##/##'
                        sheet1.cell(row, col).alignment = Alignment(wrapText=True, vertical='center', horizontal='center')
                        sheet1.cell(row, col).font = Font(size=9)
                        sheet1.cell(row, col).border = Border(bottom=Side(style='thin', color='D4D4D4'))
                        col += 1
                        # Top/Right Hinge
                        if door[10] == 0:
                            hinge_center = ''
                        elif door[10] == '-':
                            hinge_center = door[10]
                        else:
                            hinge_center = mm_to_in(door[10])
                        sheet1.cell(row, col, hinge_center)
                        sheet1.cell(row, col).number_format = '#  ##/##'
                        sheet1.cell(row, col).alignment = Alignment(wrapText=True, vertical='center', horizontal='center')
                        sheet1.cell(row, col).font = Font(size=9)
                        sheet1.cell(row, col).border = Border(bottom=Side(style='thin', color='D4D4D4'), right=Side(style='thin', color='D4D4D4'))
                        col += 1
                        
                        # convert list of cab_nums to string
                        cab_num_string = ', '.join(door[11])
                        
                        # Cabinet Numbers
                        sheet1.cell(row, col, cab_num_string)
                        sheet1.cell(row, col).alignment = Alignment(wrapText=True, vertical='center', horizontal='center')
                        sheet1.cell(row, col).font = Font(size=9)
                        sheet1.cell(row, col).border = Border(bottom=Side(style='thin', color='D4D4D4'))
                        

                        # increment quantity
                        type_total += door[0]

                        # reset row & column
                        col -= 11
                        row += 1

                    row -= 1
                    sheet1.cell(row, col).border = Border(bottom=Side(style='thin', color='000000'))
                    sheet1.cell(row, col + 1).border = Border(bottom=Side(style='thin', color='000000'))
                    sheet1.cell(row, col + 2).border = Border(bottom=Side(style='thin', color='000000'))
                    sheet1.cell(row, col + 3).border = Border(bottom=Side(style='thin', color='000000'))
                    sheet1.cell(row, col + 4).border = Border(bottom=Side(style='thin', color='000000'))
                    sheet1.cell(row, col + 5).border = Border(bottom=Side(style='thin', color='000000'))
                    sheet1.cell(row, col + 6).border = Border(bottom=Side(style='thin', color='000000'), right=Side(style='thin', color='D4D4D4'))
                    sheet1.cell(row, col + 7).border = Border(bottom=Side(style='thin', color='000000'))
                    sheet1.cell(row, col + 8).border = Border(bottom=Side(style='thin', color='000000'))
                    sheet1.cell(row, col + 9).border = Border(bottom=Side(style='thin', color='000000'))
                    sheet1.cell(row, col + 10).border = Border(bottom=Side(style='thin', color='000000'), right=Side(style='thin', color='D4D4D4'))
                    sheet1.cell(row, col + 11).border = Border(bottom=Side(style='thin', color='000000'))
                    row += 1
                    sheet1.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col + 11)

                    hinge_total_string = "Hinge Type Total: (" if _type != '' else "Total Doors w/o Hinges: ("
                    
                    sheet1.cell(row, col, hinge_total_string + str(type_total) + ")")
                    sheet1.cell(row, col).alignment = Alignment(wrapText=True, vertical='top', horizontal='left')
                    sheet1.cell(row, col).font = Font(size=10, italic=True)

                    style_total += type_total
                    row += 2
            
                sheet1.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col + 11)
                sheet1.cell(row, col, "Door Style Total: (" + str(style_total) + ")")
                sheet1.cell(row, col).alignment = Alignment(wrapText=True, vertical='center', horizontal='left')
                sheet1.cell(row, col).font = Font(size=11, bold=True, italic=True, underline='single')
                col -= 1

                mat_total += style_total
                row += 2
                
            sheet1.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col + 11)
            sheet1.cell(row, col, "Material Total: (" + str(mat_total) + ")")
            sheet1.cell(row, col).alignment = Alignment(wrapText=True, vertical='center', horizontal='left')
            sheet1.cell(row, col).font = Font(size=14, italic=True, bold=True)
            page_break = Break(id=row)
            sheet1.row_breaks.append(page_break)
            row += 1
                    
        
        sheet1.column_dimensions['A'].width = 2
        sheet1.column_dimensions['B'].width = 4
        sheet1.column_dimensions['C'].width = 11
        sheet1.column_dimensions['D'].width = 20
        sheet1.column_dimensions['E'].width = 9
        sheet1.column_dimensions['F'].width = 9
        sheet1.column_dimensions['G'].width = 4
        sheet1.column_dimensions['H'].width = 6
        sheet1.column_dimensions['I'].width = 8
        sheet1.column_dimensions['J'].width = 8
        sheet1.column_dimensions['K'].width = 8
        sheet1.column_dimensions['L'].width = 8
        sheet1.column_dimensions['M'].width = 20

        sheet1.page_margins.left = 0.5
        sheet1.page_margins.right = 0.5
        sheet1.page_margins.top = 1.0
        sheet1.page_margins.bottom = 0.5
        sheet1.page_margins.footer = 0.25
        sheet1.page_margins.header = 0.375


        sheet1.oddHeader.left.text = job_name + ' - Boring List (Rooms: ' + rooms_string + ')'
        sheet1.oddHeader.left.size = 12
        sheet1.oddHeader.left.color = "000000"
        sheet1.oddFooter.right.text = "Page &[Page] of &N"
        sheet1.oddFooter.right.size = 10
        sheet1.oddFooter.right.color = "000000"
        
        print_area = 'A1:M' + str(row - 1)
        sheet1.print_area = print_area
        sheet1.sheet_properties.pageSetUpPr.fitToPage = True
        sheet1.page_setup.fitToHeight = False   
        
        save_name = job_name + ' - Boring List - ' + rooms_string + " - " + now_string + '.xlsx'
        full_save_name = os.path.join(dir_path, save_name)
        try:
            wb.save(full_save_name)
        except PermissionError:
            print("\nSAVE FAILED\nYou will need to close the open file before it can be saved.")

        os.startfile(full_save_name)

    create_excel()

    
hinge_boring_report()


input("Press Enter to Close")
