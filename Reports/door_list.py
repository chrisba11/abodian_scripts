import os
from openpyxl.workbook import Workbook
from openpyxl.styles import Border, Side, Alignment, Font
from openpyxl.worksheet.pagebreak import Break


# This is to ask for the directory path at the command prompt
dir_path = input('What is the full directory path? ')


def door_list_report():
    """
    Generates an Excel file listing the doors separated by door style and 
        grouped by door material used for a specified job.
    The function opens each of the .des files inside a job directory and looks 
        at every product in the room.
    If the product has a door material template that isn't in the dict,
        it adds that door material template to the dictionary of mat temps.
    If the product has a door style that isn't in the dict for that mat temp,
        it adds that door style to a dictionary of styles in that mat temp.
    Then it writes data from the dictionary of styles/mat temps to an Excel file
        inside the directory it was told to run in.
    It formats the cells as it goes and sets the print area the content inserted.
    Finally, it opens the Excel file so the user can print.
    """
    dir_path_lst = dir_path.split('\\')
    job_name = dir_path_lst[-1]

    temp_dict = {}

    # temp_dict sample
    # temp_dict = {
    #     "02 - PAN - Oxford White": {
    #         "Slab Door - VG": [
    #             ["Door(L)", 555.6, 720.7, "R3C1"],
    #             ["Door(L)", 555.6, 720.7, "R3C3"],
    #             ["Door(L)", 555.6, 720.7, "R6C1"],
    #             ["Door(L)", 555.6, 720.7, "R6C3"],
    #         ],
    #         "Slab Dw - VG": [
    #             ["Door(L)", 479.4, 281, "R3C2"],
    #             ["Door(L)", 479.4, 281, "R6C2"],
    #             ["Door(L)", 479.4, 280.9, "R3C2"],
    #             ["Door(L)", 479.4, 280.9, "R6C2"],
    #         ]
    #     },
    #     "02 - 5pc Painted - In House MDF (2)": {
    #         "Shaker Door": [
    #             ["Door(L)", 454, 771.5, "R1C3"],
    #             ["Door(R)", 454, 771.5, "R1C3"],
    #             ["Door(L)", 454, 771.5, "R4C5"],
    #             ["Door(R)", 454, 771.5, "R4C5"]
    #         ],
    #         "Shaker DF Top": [
    #             ["Drawer", 936.6, 152.4, "R1C2"],
    #             ["Drawer", 936.6, 152.4, "R4C6"],
    #         ],
    #         "Shaker DF Mid/Bot": [
    #             ["Drawer", 936.6, 306.4, "R1C2"],
    #             ["Drawer", 936.6, 306.3, "R1C2"],
    #             ["Drawer", 936.6, 306.4, "R4C6"],
    #             ["Drawer", 936.6, 306.3, "R4C6"],
    #         ]
    #     },
    #     "02 - PAN - Absolute Acajou": {
    #         "Slab Door - VG": [
    #             ["Door(L)", 555.6, 720.7, "R3C1"],
    #             ["Door(L)", 555.6, 720.7, "R3C3"],
    #             ["Door(L)", 555.6, 720.7, "R6C1"],
    #             ["Door(L)", 555.6, 720.7, "R6C3"],
    #         ],
    #         "Slab Dw - VG": [
    #             ["Door(L)", 479.4, 281, "R3C2"],
    #             ["Door(L)", 479.4, 281, "R6C2"],
    #             ["Door(L)", 479.4, 280.9, "R3C2"],
    #             ["Door(L)", 479.4, 280.9, "R6C2"],
    #         ]
    #     }
    # }
    
    for root, dirs, files in os.walk(dir_path):
        
        for file in files:

            if file.endswith('.des'):

                full_path = dir_path + '\\' + file
                f = open(full_path, "rt")
                content = f.readlines()
                f.close()

                # room settings
                room_num = file[4:file.find('.')]
                mat_temp_start_idx = content[4].find('MatDoorTemplate=') + 17
                mat_temp_end_idx = content[4].find('" MatDrawerTemplate=')
                room_mat_temp = content[4][mat_temp_start_idx:mat_temp_end_idx]

                if room_mat_temp not in temp_dict:
                    temp_dict[room_mat_temp] = {}

                # declaring variables that are used in multiple scopes
                numbered = ''
                door_mat_or = ''
                door_style = ''
                report_name = ''
                width = 0
                height = 0

    
                for line in content:
                    if line.startswith('    <Product '):
                        # product number
                        num_start_idx = line.find('CabNo=') + 7
                        num_end_idx = line.find(' Numbered=')
                        prod_num = line[num_start_idx:num_end_idx - 1]
                        is_numbered_start_idx = num_end_idx + 11
                        is_numbered = line[is_numbered_start_idx:is_numbered_start_idx + 1]
                        if is_numbered == 'T':
                            numbered = 'C'
                        else:
                            numbered = 'N'
                        # product material overrides
                        # door mat
                        mat_or_start_idx = line.find('MatDoorOR=') + 11
                        mat_or_end_idx = line.find('" MatDrawerOR=')
                        door_mat_or = line[mat_or_start_idx:mat_or_end_idx]



                    if line.startswith('        <ProductDoor '):
                        # door style
                        start_idx = line.find('DoorStyle=') + 11
                        end_idx = line.find('" W=')
                        door_style = line[start_idx:end_idx]
                        # width
                        start_idx = end_idx + 5
                        end_idx = line.find('" H=')
                        width = float(line[start_idx:end_idx])
                        # height
                        start_idx = end_idx + 5
                        end_idx = line.find('" Oversize=')
                        height = float(line[start_idx:end_idx])

                    if line.startswith('          <DrawerFront '):
                        # door style
                        start_idx = line.find('DoorStyle=') + 11
                        end_idx = line.find('" W=')
                        door_style = line[start_idx:end_idx]
                        # width
                        start_idx = end_idx + 5
                        end_idx = line.find('" H=')
                        width = float(line[start_idx:end_idx])
                        # height
                        start_idx = end_idx + 5
                        end_idx = line.find('" Oversize=')
                        height = float(line[start_idx:end_idx])
                        
                    if line.startswith('          <DoorProdPart ') or line.startswith('            <DoorProdPart '):
                        start_idx = line.find('ReportName=') + 12
                        end_idx = line.find('" UsageType=')
                        report_name = line[start_idx:end_idx]

                        # append to dict
                        if door_mat_or == '':
                            if door_style in temp_dict[room_mat_temp]:
                                temp_dict[room_mat_temp][door_style].append(
                                    [report_name, width, height, room_num, numbered, prod_num]
                                )
                            else:
                                temp_dict[room_mat_temp][door_style] = [
                                    [report_name, width, height, room_num, numbered, prod_num]
                                ]
                        elif door_mat_or in temp_dict:
                            if door_style in temp_dict[door_mat_or]:
                                temp_dict[door_mat_or][door_style].append(
                                    [report_name, width, height, room_num, numbered, prod_num]
                                    )
                            else:
                                temp_dict[door_mat_or][door_style] = [
                                    [report_name, width, height, room_num, numbered, prod_num]
                                ]
                        else:
                            temp_dict[door_mat_or] = {
                                door_style: [
                                    [report_name, width, height, room_num, numbered, prod_num]
                                ]
                            }
    
    for template in temp_dict:
        for door_style in temp_dict[template]:
            temp_dict[template][door_style].sort(key=lambda x: int(x[5]))
            temp_dict[template][door_style].sort(key=lambda x: int(x[3]))
            temp_dict[template][door_style].reverse()
            temp_dict[template][door_style].sort(key=lambda x: x[2])
            temp_dict[template][door_style].sort(key=lambda x: x[1])
            temp_dict[template][door_style].reverse()

    # print('--------')        
    # for template in temp_dict:
    #     if temp_dict[template]:
    #         print('Template Name:', template)
    #         for door_style in temp_dict[template]:
    #             print('   Door Style:', door_style)
    #             for door in temp_dict[template][door_style]:
    #                 print('     ', door[0] + ',', str(door[1]) + ',', str(door[2]) + ',', door[3])
    # print('--------')


    wb = Workbook()
    sheet1 = wb.active
    row = 1
    col = 1

    for template in temp_dict:
        if temp_dict[template]:

            sheet1.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col + 3)
            sheet1.cell(row, col, job_name + ' - Door List')
            sheet1.cell(row, col).alignment = Alignment(vertical='top')
            sheet1.cell(row, col).font = Font(size=12, bold=True, italic=True)
            sheet1.row_dimensions[row].height = 40
            row += 1

            sheet1.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col + 3)
            sheet1.cell(row, col, template)
            sheet1.cell(row, col).alignment = Alignment(wrapText=True, horizontal='left')
            sheet1.cell(row, col).font = Font(size=14, bold=True, underline='single')
            row += 1


            for door_style in temp_dict[template]:
                sheet1.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col + 3)
                sheet1.cell(row, col, door_style)
                sheet1.cell(row, col).alignment = Alignment(wrapText=True, vertical='top', indent=1.0)
                sheet1.cell(row, col).font = Font(size=12, bold=True, italic=True)
                sheet1.cell(row, col).border = Border(bottom=Side(style='thin', color='D4D4D4'))
                sheet1.cell(row, col + 1).border = Border(bottom=Side(style='thin', color='D4D4D4'))
                sheet1.cell(row, col + 2).border = Border(bottom=Side(style='thin', color='D4D4D4'))
                sheet1.cell(row, col + 3).border = Border(bottom=Side(style='thin', color='D4D4D4'))
                row += 1


                for door in temp_dict[template][door_style]:
                    sheet1.cell(row, col, door[0])
                    sheet1.cell(row, col).alignment = Alignment(wrapText=True, vertical='top', horizontal='left', indent=4.0)
                    sheet1.cell(row, col).font = Font(size=11)
                    sheet1.cell(row, col).border = Border(bottom=Side(style='thin', color='D4D4D4'))
                    col += 1
                    sheet1.cell(row, col, round(door[1],1))
                    sheet1.cell(row, col).alignment = Alignment(wrapText=True, vertical='top', horizontal='left', indent=1.0)
                    sheet1.cell(row, col).font = Font(size=11)
                    sheet1.cell(row, col).border = Border(bottom=Side(style='thin', color='D4D4D4'))
                    col += 1
                    sheet1.cell(row, col, round(door[2],1))
                    sheet1.cell(row, col).alignment = Alignment(wrapText=True, vertical='top', horizontal='left', indent=1.0)
                    sheet1.cell(row, col).font = Font(size=11)
                    sheet1.cell(row, col).border = Border(bottom=Side(style='thin', color='D4D4D4'))
                    col += 1
                    sheet1.cell(row, col, 'R' + door[3] + door[4] + door[5])
                    sheet1.cell(row, col).alignment = Alignment(wrapText=True, vertical='top', horizontal='left', indent=1.0)
                    sheet1.cell(row, col).font = Font(size=11)
                    sheet1.cell(row, col).border = Border(bottom=Side(style='thin', color='D4D4D4'))
                    col -= 3
                    row += 1
        
            page_break = Break(id=row)
            sheet1.row_breaks.append(page_break)
            row += 1


    sheet1.column_dimensions['A'].width = 20
    sheet1.column_dimensions['B'].width = 10
    sheet1.column_dimensions['C'].width = 10
    sheet1.column_dimensions['D'].width = 40

    sheet1.oddFooter.right.text = "Page &[Page] of &N"
    sheet1.oddFooter.right.size = 9
    sheet1.oddFooter.right.color = "000000"
    
    print_area = 'A1:D' + str(row - 1)
    sheet1.print_area = print_area
    
    save_name = job_name + ' - Door List' + '.xlsx'
    full_save_name = os.path.join(dir_path, save_name)
    try:
        wb.save(full_save_name)
    except PermissionError:
        print("\nSAVE FAILED\nYou will need to close the open file before it can be saved.")

    os.startfile(full_save_name)
    
door_list_report()
