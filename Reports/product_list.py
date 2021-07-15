import os
from openpyxl.workbook import Workbook
from openpyxl.styles import Border, Side, Alignment, Font


# This is to ask for the directory path at the command prompt
dir_path = input('What is the full directory path? ')


def product_list():
    """
    Generates an Excel file listing all of the products in the job that have special notes.
    The function opens each of the .des files inside a job directory and looks 
        at every product in the room.
    If the product has notes, it adds that product to a dictionary of products.
    Then it writes data from the dictionary of products to an Excel file
        inside the directory it was told to run in.
    It formats the cells as it goes and sets the print area the content inserted.
    Finally, it opens the Excel file so the user can print.
    """
    dir_path_lst = dir_path.split('\\')
    job_name = dir_path_lst[-1]

    xml_char_ents = [
        ['&quot;', '"'],
        ['&#xD;&#xA;', '\n'],
        ['&amp;', '&'],
        ['&apos;', "'"],
        ['&gt;', '>'],
        ['&lt;', '<']
    ]

    prod_dict = {}
    
    for root, dirs, files in os.walk(dir_path):
        
        for file in files:

            if file.endswith('.des'):

                full_path = dir_path + '\\' + file
                f = open(full_path, "rt")
                content = f.readlines()
                rm_start_idx = content[2].find('Name=') + 6
                rm_end_idx = content[2].find('" RoomNosDirty=')
                room_num = int(file[4:file.find('.')])
                room_name = content[2][rm_start_idx:rm_end_idx] + ' (Room' + str(room_num) + ')'
                prod_dict[room_num] = [room_name,[]]
                                
                for line in content:
                    if line.startswith('    <Product '):
                        note_start_idx = line.find('Notes=') + 6
                        note_end_idx = line.find('" Price=')
                        note_intro = line[note_start_idx:note_start_idx + 2]

                        if note_intro != '""':
                            prod_start_idx = line.find('ProdName=') + 10
                            prod_end_idx = line.find('" IDTag=')
                            prod_name = line[prod_start_idx:prod_end_idx]
                            num_start_idx = line.find('CabNo=') + 7
                            num_end_idx = line.find(' Numbered=')
                            prod_num = line[num_start_idx:num_end_idx - 1]
                            is_numbered_start_idx = num_end_idx + 11
                            is_numbered = line[is_numbered_start_idx:is_numbered_start_idx + 1]
                            note = line[note_start_idx + 1:note_end_idx]

                            if is_numbered == 'T':
                                full_prod_num = 'R' + str(room_num) + 'C' + prod_num
                            else:
                                full_prod_num = 'R' + str(room_num) + 'N' + prod_num
                                
                            for char in xml_char_ents:
                                prod_name = prod_name.replace(char[0], char[1])
                                note = note.replace(char[0], char[1])
                            
                            prod_dict[room_num][1].append([prod_name, full_prod_num, note])


    wb = Workbook()
    sheet1 = wb.active
    row = 1
    col = 1
    room_key = 0

    sheet1.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col + 2)
    sheet1.cell(row, col, job_name + ' - Special Notes')
    sheet1.cell(row, col).alignment = Alignment(vertical='top')
    sheet1.cell(row, col).font = Font(size=12, bold=True, italic=True)
    sheet1.row_dimensions[1].height = 25
    row += 1

    for i in range(len(prod_dict)):
        if prod_dict[room_key][1] != []:
            sheet1.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col + 2)
            sheet1.cell(row, col, prod_dict[room_key][0])
            sheet1.cell(row, col).alignment = Alignment(wrapText=True, horizontal='center')
            sheet1.cell(row, col).font = Font(size=14, bold=True, underline='single')
            row += 1

            sheet1.cell(row, col, 'Product Name')
            sheet1.cell(row, col).alignment = Alignment(horizontal='general', indent=1.0)
            sheet1.cell(row, col).font = Font(size=12, italic=True)
            sheet1.cell(row, col).border = Border(bottom=Side(style='thin', color='000000'))
            col += 1

            sheet1.cell(row, col, 'CabNo')
            sheet1.cell(row, col).alignment = Alignment(horizontal='center')
            sheet1.cell(row, col).font = Font(size=12, italic=True)
            sheet1.cell(row, col).border = Border(bottom=Side(style='thin', color='000000'))
            col += 1

            sheet1.cell(row, col, 'Notes')
            sheet1.cell(row, col).alignment = Alignment(horizontal='general', indent=1.0)
            sheet1.cell(row, col).font = Font(size=12, italic=True)
            sheet1.cell(row, col).border = Border(bottom=Side(style='thin', color='000000'))
            col -= 2
            row += 1


            for prod in prod_dict[room_key][1]:
                sheet1.cell(row, col, prod[0])
                sheet1.cell(row, col).alignment = Alignment(wrapText=True, vertical='top', indent=1.0)
                sheet1.cell(row, col).font = Font(size=10)
                sheet1.cell(row, col).border = Border(bottom=Side(style='thin', color='D4D4D4'))
                col += 1

                sheet1.cell(row, col, prod[1])
                sheet1.cell(row, col).alignment = Alignment(wrapText=True, vertical='top', horizontal='center')
                sheet1.cell(row, col).font = Font(size=10)
                sheet1.cell(row, col).border = Border(bottom=Side(style='thin', color='D4D4D4'))
                col += 1
                
                sheet1.cell(row, col, prod[2])
                sheet1.cell(row, col).alignment = Alignment(wrapText=True, vertical='top', indent=1.0)
                sheet1.cell(row, col).font = Font(size=10)
                sheet1.cell(row, col).border = Border(bottom=Side(style='thin', color='D4D4D4'))
                col -= 2
                row += 1

            row += 1
        room_key += 1

    sheet1.column_dimensions['A'].width = 25
    sheet1.column_dimensions['B'].width = 8
    sheet1.column_dimensions['C'].width = 50
    
    print_area = 'A1:C' + str(row)
    sheet1.print_area = print_area
    
    save_name = job_name + ' - Special Notes' + '.xlsx'
    full_save_name = os.path.join(dir_path, save_name)
    try:
        wb.save(full_save_name)
    except PermissionError:
        print("\nSAVE FAILED\nYou will need to close the open file before it can be saved.")

    os.startfile(full_save_name)
    
product_list()

input("Press Enter to Close")