import os
from openpyxl.workbook import Workbook
from openpyxl.styles import Border, Side, Alignment, Font
from datetime import datetime


# This is to ask for the directory path at the command prompt
dir_path = input('What is the full directory path? ')
rooms_string = input('Which rooms should be included in this report? List the room integers only, separated by commas and no spaces. Include "0" if you want Order Entry included. If all rooms, type "all". ').lower()
rooms_in_report = rooms_string.split(',')

now = datetime.now()
now_string = now.strftime("%m.%d.%Y-%H.%M.%S")


def inserts_list():
    """
    """
    dir_path_lst = dir_path.split('\\')
    job_name = dir_path_lst[-1]

    xml_char_ents = [
        ['&quot;', '"'],
        ['&#xD;&#xA;', '\n'],
        ['&amp;', '&'],
        ['&apos;', "'"],
        ['&gt;', '>'],
        ['&lt;', '<']
    ]

    prod_dict = {}
    
    for root, dirs, files in os.walk(dir_path):
        
        for file in files:

            if file.endswith('.des'):

                full_path = dir_path + '\\' + file
                f = open(full_path, "rt")
                content = f.readlines()
                rm_start_idx = content[2].find('Name=') + 6
                rm_end_idx = content[2].find('" RoomNosDirty=')
                room_num = int(file[4:file.find('.')])
                if room_num not in rooms_in_report and 'all' not in rooms_in_report:
                    continue
                room_name = content[2][rm_start_idx:rm_end_idx] + ' (Room' + str(room_num) + ')'
                prod_dict[room_num] = [room_name,[]]
                                
                for line in content:
                    if line.startswith('    <Product '):
                        prod_start_idx = line.find('ProdName=') + 10
                        prod_end_idx = line.find('" IDTag=')
                        prod_name = line[prod_start_idx:prod_end_idx]
                        num_start_idx = line.find('CabNo=') + 7
                        num_end_idx = line.find(' Numbered=')
                        prod_num = line[num_start_idx:num_end_idx - 1]
                        is_numbered_start_idx = num_end_idx + 11
                        is_numbered = line[is_numbered_start_idx:is_numbered_start_idx + 1]
                        inserts =[]


                        if is_numbered == 'T':
                            full_prod_num = 'R' + str(room_num) + 'C' + prod_num
                        else:
                            full_prod_num = 'R' + str(room_num) + 'N' + prod_num
                            

                        
                    if line.startswith('          <Insert'):
                        if 'GraphicOnly="False"' in line:
                            insert_start_idx = line.find('Name=') + 5
                            insert_end_idx = line.find('" Library=')
                            insert_name = line[insert_start_idx + 1:insert_end_idx]

                            opening_start_idx = line.find('<Insert') + 7
                            opening_end_idx = line.find('Count=')
                            opening_name = line[opening_start_idx:opening_end_idx]

                            for char in xml_char_ents:
                                prod_name = prod_name.replace(char[0], char[1])
                                insert_name = insert_name.replace(char[0], char[1])

                            inserts.append([insert_name, opening_name])

                                              
                    if line.startswith('    </Product>'):
                        if len(inserts) > 0:                
                            prod_dict[room_num][1].append([full_prod_num, prod_name, inserts])

    if len(prod_dict) > 0:
        wb = Workbook()
        sheet1 = wb.active
        row = 1
        col = 1
        room_key = 0

        for i in range(len(prod_dict)):
            if prod_dict[room_key][1] != []:
                sheet1.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col + 3)
                sheet1.cell(row, col, prod_dict[room_key][0])
                sheet1.cell(row, col).alignment = Alignment(wrapText=True, horizontal='center')
                sheet1.cell(row, col).font = Font(size=14, bold=True, underline='single')
                row += 1

                sheet1.cell(row, col, 'Product Name')
                sheet1.cell(row, col).alignment = Alignment(horizontal='general', indent=1.0)
                sheet1.cell(row, col).font = Font(size=12, italic=True)
                sheet1.cell(row, col).border = Border(bottom=Side(style='thin', color='000000'))
                col += 1

                sheet1.cell(row, col, 'Insert Name')
                sheet1.cell(row, col).alignment = Alignment(horizontal='general', indent=1.0)
                sheet1.cell(row, col).font = Font(size=12, italic=True)
                sheet1.cell(row, col).border = Border(bottom=Side(style='thin', color='000000'))
                col += 1

                sheet1.cell(row, col, 'Opening')
                sheet1.cell(row, col).alignment = Alignment(horizontal='center')
                sheet1.cell(row, col).font = Font(size=12, italic=True)
                sheet1.cell(row, col).border = Border(bottom=Side(style='thin', color='000000'))
                col += 1

                sheet1.cell(row, col, 'CabNo')
                sheet1.cell(row, col).alignment = Alignment(horizontal='center')
                sheet1.cell(row, col).font = Font(size=12, italic=True)
                sheet1.cell(row, col).border = Border(bottom=Side(style='thin', color='000000'))
                col -= 3
                row += 1


                for prod in prod_dict[room_key][1]:
                    for insert in prod[2]:
                        sheet1.cell(row, col, prod[1])
                        sheet1.cell(row, col).alignment = Alignment(wrapText=True, vertical='top', indent=1.0)
                        sheet1.cell(row, col).font = Font(size=10)
                        sheet1.cell(row, col).border = Border(bottom=Side(style='thin', color='D4D4D4'))
                        col += 1

                        sheet1.cell(row, col, insert[0])
                        sheet1.cell(row, col).alignment = Alignment(wrapText=True, vertical='top', indent=1.0)
                        sheet1.cell(row, col).font = Font(size=10)
                        sheet1.cell(row, col).border = Border(bottom=Side(style='thin', color='D4D4D4'))
                        col += 1

                        sheet1.cell(row, col, insert[1])
                        sheet1.cell(row, col).alignment = Alignment(wrapText=True, vertical='top', horizontal='center')
                        sheet1.cell(row, col).font = Font(size=10)
                        sheet1.cell(row, col).border = Border(bottom=Side(style='thin', color='D4D4D4'))
                        col += 1
                        
                        sheet1.cell(row, col, prod[0])
                        sheet1.cell(row, col).alignment = Alignment(wrapText=True, vertical='top', horizontal='center')
                        sheet1.cell(row, col).font = Font(size=10)
                        sheet1.cell(row, col).border = Border(bottom=Side(style='thin', color='D4D4D4'))
                        col -= 3
                        row += 1

                row += 1
            room_key += 1

        sheet1.column_dimensions['A'].width = 25
        sheet1.column_dimensions['B'].width = 55
        sheet1.column_dimensions['C'].width = 9
        sheet1.column_dimensions['D'].width = 8
        
        sheet1.page_margins.left = 0.5
        sheet1.page_margins.right = 0.5
        sheet1.page_margins.top = 1.0
        sheet1.page_margins.bottom = 0.5
        sheet1.page_margins.footer = 0.25
        sheet1.page_margins.header = 0.375


        sheet1.oddHeader.left.text = job_name + ' - Insert List (Rooms: ' + rooms_string + ')'
        sheet1.oddHeader.left.size = 12
        sheet1.oddHeader.left.color = "000000"
        sheet1.oddFooter.right.text = "Page &[Page] of &N"
        sheet1.oddFooter.right.size = 10
        sheet1.oddFooter.right.color = "000000"

        print_area = 'A1:D' + str(row)
        sheet1.print_area = print_area
        sheet1.sheet_properties.pageSetUpPr.fitToPage = True
        sheet1.page_setup.fitToHeight = False   

        
        save_name = job_name + ' - Insert List'  + rooms_string + " - " + now_string + '.xlsx'
        full_save_name = os.path.join(dir_path, save_name)
        try:
            wb.save(full_save_name)
        except PermissionError:
            print("\nSAVE FAILED\nYou will need to close the open file before it can be saved.")

        os.startfile(full_save_name)
    
    else:
        print("\nThere are no inserts in the rooms you specified.")
    
inserts_list()

input("\nPress Enter to Close")