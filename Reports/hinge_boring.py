import os
import pprint
from openpyxl.workbook import Workbook
from openpyxl.styles import Border, Side, Alignment, Font
from openpyxl.worksheet.pagebreak import Break


# This is to ask for the directory path at the command prompt
dir_path = input('What is the full directory path for the job? ')
rooms_in_report = input('Which rooms should be included in this report? List the room integers only, separated by commas and no spaces. Include "0" if you want Order Entry included. If all rooms, type "all". ').lower()


def hinge_boring_report():
    """

    """
    dir_path_lst = dir_path.split('\\')
    job_name = dir_path_lst[-1]

    product_dict = {}

# #    product_dict sample
#     product_dict = {
#         "Room1": {
#             "MatDoorTemplate": "02 - 5pc Painted - In House MDF (1)",
#             "Products": {
#                 "C4": {
#                     "UniqueID": "3101352",
#                     "Name": "Base Sink",
#                     "MatOR": None,
#                     "Doors": [
#                         {
#                             "DoorStyle": "Shaker Door",
#                             "Name": "Door(L)",
#                             "ReportName": "Door(L)",
#                             "Comment": "Cabinet Door",
#                             "Quan": 1,
#                             "W": 454.0125,
#                             "H": 771.5,
#                             "HingeCenterLines": [101.6, 542.9],
#                             "HingeEdge": "Left",
#                             "HingeType": "DTC C-80 110 / 2mm - 105-C80A675NF",
#                             "IsHorizGrain": False,
#                         },
#                         {
#                             "DoorStyle": "Shaker Door",
#                             "Name": "Door(R)",
#                             "ReportName": "Door(R)",
#                             "Comment": "Cabinet Door",
#                             "Quan": 1,
#                             "W": 454.0125,
#                             "H": 771.5,
#                             "HingeCenterLines": [101.6, 542.9],
#                             "HingeEdge": "Right",
#                             "HingeType": "DTC C-80 110 / 2mm - 105-C80A675NF",
#                             "IsHorizGrain": False,
#                         },
#                     ]
#                 },
#                 "C6": {
#                     "UniqueID": "3101389",
#                     "Name": "Base Corner Left - Susan (L)",
#                     "MatOR": None,
#                     "Doors": [
#                         {
#                             "DoorStyle": "Shaker Door",
#                             "Name": "Door(L)",
#                             "ReportName": "Door(L)",
#                             "Comment": "Cabinet Door",
#                             "Quan": 1,
#                             "W": 454.0125,
#                             "H": 771.5,
#                             "HingeCenterLines": [101.6, 542.9],
#                             "HingeEdge": "Left",
#                             "HingeType": "DTC C-80 110 / 2mm - 105-C80A675NF",
#                             "Horizontal": False,
#                         },
#                         {
#                             "DoorStyle": "Shaker Door",
#                             "Name": "Door(L)",
#                             "ReportName": "Door(L)",
#                             "Comment": "Cabinet Door",
#                             "Quan": 1,
#                             "W": 454.0125,
#                             "H": 771.5,
#                             "HingeCenterLines": [101.6, 542.9],
#                             "HingeEdge": "Right",
#                             "HingeType": "DTC C-80 110 / 2mm - 105-C80A675NF",
#                             "Horizontal": False,
#                         },
#                     ]
#                 },
#             },
#         },
#     }

    
    for root, dirs, files in os.walk(dir_path):
        
        for file in files:

            if file.endswith('.des'):

                full_path = dir_path + '\\' + file
                f = open(full_path, "rt")
                content = f.readlines()
                f.close()

                # room settings
                room_num = file[4:file.find('.')]
                # only record rooms requested for this report
                if room_num not in rooms_in_report and 'all' not in rooms_in_report:
                    continue
                mat_temp_start_idx = content[4].find('MatDoorTemplate=') + 17
                mat_temp_end_idx = content[4].find('" MatDrawerTemplate=')
                room_mat_door_template = content[4][mat_temp_start_idx:mat_temp_end_idx]

                # add room to product list dictionary
                product_dict[room_num] = {
                    "MatDoorTemplate": room_mat_door_template,
                    "Products": {}
                }

                # set products var to current room's products list
                products = product_dict[room_num]["Products"]

                # declaring variables that are used in multiple scopes
                doors = None
                door = None

                for line in content:
                    if line.startswith('    <Product '):
                        # product number
                        start_idx = line.find('CabNo=') + 7
                        end_idx = line.find('" Numbered=')
                        prod_num = line[start_idx:end_idx]
                        
                        # product number prefix
                        start_idx = end_idx + 12
                        num_prefix = 'C' if line[start_idx:start_idx + 1] == 'T' else 'N'
                        
                        # unique id
                        start_idx = line.find('UniqueID=') + 10
                        end_idx = line.find('" OrderID=')
                        unique_id = line[start_idx:end_idx]

                        # product name
                        start_idx = line.find('ProdName=') + 10
                        end_idx = line.find('" IDTag=')
                        prod_name = line[start_idx:end_idx]

                        # door material override
                        start_idx = line.find('MatDoorOR=') + 11
                        end_idx = line.find('" MatDrawerOR=')
                        door_mat_or = line[start_idx:end_idx] if line[start_idx:end_idx] != '' else None

                        # add product to room's product list
                        products[num_prefix + prod_num] = {
                            "UniqueID": unique_id,
                            "Name": prod_name,
                            "MatOR": door_mat_or,
                            "Doors": []
                        }

                        # set doors var to current product's door list
                        doors = products[num_prefix + prod_num]["Doors"]


                    if line.startswith('        <ProductDoor '):
                        door = {}

                        # door style
                        start_idx = line.find('DoorStyle=') + 11
                        end_idx = line.find('" W=')
                        door_style = line[start_idx:end_idx]

                        # width
                        start_idx = end_idx + 5
                        end_idx = line.find('" H=')
                        width = float(line[start_idx:end_idx])
                        
                        # height
                        start_idx = end_idx + 5
                        end_idx = line.find('" Oversize=')
                        height = float(line[start_idx:end_idx])

                        # hinge center lines
                        start_idx = line.find('HingeCenterLines=') + 18
                        end_idx = line.find('" HingeEdge=')
                        hinge_centers = line[start_idx:end_idx].split('#')
                        hinge_centers = [float(i) for i in hinge_centers if i != '']

                        # hinge edge
                        start_idx = end_idx + 13
                        end_idx = line.find('" HingeType=')
                        hinge_edge = line[start_idx:end_idx]

                        # hinge type
                        start_idx = end_idx + 13
                        end_idx = line.find('" IsDrawerFront=')
                        hinge_type = line[start_idx:end_idx]

                        # horizontal grain?
                        start_idx = line.find('IsHorizGrain=') + 14
                        end_idx = line.find('" DoorType=')
                        horizontal = line[start_idx:end_idx]
                        horizontal = True if horizontal == 'True' else False

                        # add to door dict
                        door["DoorStyle"] = door_style
                        door["W"] = width
                        door["H"] = height
                        door["HingeCenterLines"] = hinge_centers
                        door["HingeEdge"] = hinge_edge
                        door["HingeType"] = hinge_type
                        door["IsHorizGrain"] = horizontal


                    if line.startswith('          <DoorProdPart '):
                        # door name
                        start_idx = line.find('Name=') + 6
                        end_idx = line.find('" ReportName=')
                        door_name = line[start_idx:end_idx]
                        
                        # report name
                        start_idx = end_idx + 14
                        end_idx = line.find('" UsageType=')
                        report_name = line[start_idx:end_idx]

                        # comment
                        start_idx = line.find('Comment=') + 9
                        end_idx = line.find('" CommentLocked=')
                        comment = line[start_idx:end_idx]

                        # quantity
                        start_idx = line.find('Quan=') + 6
                        end_idx = line.find('" W=')
                        quantity = line[start_idx:end_idx]

                        # add to door dict
                        door["Name"] = door_name
                        door["ReportName"] = report_name
                        door["Comment"] = comment
                        door["Quan"] = int(quantity)


                    if line.startswith('        </ProductDoor>'):
                        doors.append(door)


    sorted_product_dict = {}
    materials = set()
    door_styles = set()
    hinge_types = set()
    std_hinge_dist = 101.6

    for _room in product_dict:
        __room = product_dict[_room]
        mat = __room["MatDoorTemplate"]
        # if material does not exist in sorted product list, add it
        if mat not in sorted_product_dict:
            sorted_product_dict[mat] = {}
            materials.add(mat)

        for _product in __room["Products"]:
            __product = __room["Products"][_product]
            cab_num = 'R' + _room + _product

            mat = __room["MatDoorTemplate"]
            # if product has material override
            if __product["MatOR"] != None:
                mat = __product["MatOR"]

                # if material does not exist in sorted product list, add it
                if mat not in sorted_product_dict:
                    sorted_product_dict[mat] = {}
                    materials.add(mat)
            
            for _door in __product["Doors"]:
                mat_dict = sorted_product_dict[mat]
                door_style = _door["DoorStyle"]
                hinge_type = _door["HingeType"]

                # if door style is not in material dict, add it
                if door_style not in mat_dict:
                    mat_dict[door_style] = {}
                    door_styles.add(door_style)

                # if hinge type is not in door style dict, add it
                if hinge_type not in mat_dict[door_style]:
                    mat_dict[door_style][hinge_type] = []
                    hinge_types.add(hinge_type)
                
                # simplify access to door list for this door's hinge type
                door_list = mat_dict[door_style][hinge_type]

                # assign true hinge centers from top/bottom of door for up to 4 hinges
                hinge_centers = _door["HingeCenterLines"]
                num_hinges = len(hinge_centers)
                bot_hinge_center = round(hinge_centers[0],1) if num_hinges > 0 else 0
                bot_mid_hinge_center = round(hinge_centers[1],1) if num_hinges > 2 else 0
                top_mid_hinge_center = round(hinge_centers[2],1) if num_hinges > 3 else 0
                if _door["HingeEdge"] == 'Left' or _door["HingeEdge"] == 'Right':
                    top_hinge_center = round(_door["H"] - hinge_centers[-1],1) if num_hinges > 1 else 0
                else:
                    top_hinge_center = round(_door["W"] - hinge_centers[-1],1) if num_hinges > 1 else 0


                # is the hinging standard or not
                std_not = "S"
                if bot_hinge_center != std_hinge_dist or top_hinge_center != std_hinge_dist:
                    std_not = "N"
                
                door_details = [
                    _door["Quan"],
                    _door["Name"],
                    _door["Comment"],
                    _door["W"],
                    _door["H"],
                    std_not,
                    _door["HingeEdge"],
                    bot_hinge_center,
                    bot_mid_hinge_center,
                    top_mid_hinge_center,
                    top_hinge_center,
                    [cab_num]
                ]

                # if door with matching details does not exist in the list, add it
                # if a matching door does exist, add the cabinet number to the list of cab numbers
                # and increment the corresponding quantity in the quantity list
                found_match = False
                for door in door_list:
                    if door_details[1:-1] == door[1:-1]:
                        door[0] += door_details[0]
                        for cab_no in door_details[-1]:
                            door[-1].append(cab_no)
                        found_match = True
                
                if found_match is False:
                    door_list.append(door_details)

    pprint.pprint(product_dict)
    pprint.pprint(sorted_product_dict)

    # remove any materials that don't have doors associated with them
    temp_mat_set = set()
    for _material in materials:
        for _style in door_styles:
            if _style not in sorted_product_dict[_material]:
                continue

            for _type in hinge_types:
                if _type not in sorted_product_dict[_material][_style]:
                    continue

                list_of_doors = sorted_product_dict[_material][_style][_type]
                if len(list_of_doors) > 0:
                    temp_mat_set.add(_material)
    
    # creates sorted lists from sets
    mat_list = sorted(temp_mat_set)
    styles_list = sorted(door_styles)
    types_list = sorted(hinge_types)


    wb = Workbook()
    sheet1 = wb.active
    row = 1
    col = 1

    sheet1.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col + 12)
    sheet1.cell(row, col, job_name + ' - Door List')
    sheet1.cell(row, col).alignment = Alignment(vertical='top')
    sheet1.cell(row, col).font = Font(size=12, bold=True, italic=True)
    sheet1.row_dimensions[row].height = 40
    row += 1
    
    for _material in mat_list:
        sheet1.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col + 12)
        sheet1.cell(row, col, _material)
        sheet1.cell(row, col).alignment = Alignment(wrapText=True, vertical='center', horizontal='left')
        sheet1.cell(row, col).font = Font(size=16, bold=True)
        sheet1.row_dimensions[row].height = 18
        row += 2

        mat = sorted_product_dict[_material]

        mat_total = 0

        for _style in styles_list:
            if _style in mat:
                door_style = mat[_style]
            else:
                continue

            sheet1.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col + 12)
            sheet1.cell(row, col, _style)
            sheet1.cell(row, col).alignment = Alignment(wrapText=True, horizontal='left')
            sheet1.cell(row, col).font = Font(size=14, bold=True, underline='single')
            sheet1.row_dimensions[row].height = 18
            row += 2
            col += 1

            style_total = 0

            for _type in types_list:
                if _type in door_style:
                    hinge = door_style[_type]
                else:
                    continue

                if len(hinge) == 0:
                    continue

                sheet1.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col + 11)
                sheet1.cell(row, col, _type)
                sheet1.cell(row, col).alignment = Alignment(wrapText=True, vertical='top', horizontal='left')
                sheet1.cell(row, col).font = Font(size=12, bold=True, italic=True)
                sheet1.cell(row, col).border = Border(bottom=Side(style='thin', color='000000'))
                sheet1.cell(row, col + 1).border = Border(bottom=Side(style='thin', color='000000'))
                sheet1.cell(row, col + 2).border = Border(bottom=Side(style='thin', color='000000'))
                sheet1.cell(row, col + 3).border = Border(bottom=Side(style='thin', color='000000'))
                sheet1.cell(row, col + 4).border = Border(bottom=Side(style='thin', color='000000'))
                sheet1.cell(row, col + 5).border = Border(bottom=Side(style='thin', color='000000'))
                sheet1.cell(row, col + 6).border = Border(bottom=Side(style='thin', color='000000'))
                sheet1.cell(row, col + 7).border = Border(bottom=Side(style='thin', color='000000'))
                sheet1.cell(row, col + 8).border = Border(bottom=Side(style='thin', color='000000'))
                sheet1.cell(row, col + 9).border = Border(bottom=Side(style='thin', color='000000'))
                sheet1.cell(row, col + 10).border = Border(bottom=Side(style='thin', color='000000'))
                sheet1.cell(row, col + 11).border = Border(bottom=Side(style='thin', color='000000'))
                row += 1
                sheet1.cell(row, col, "Qty")
                sheet1.cell(row, col).alignment = Alignment(wrapText=True, vertical='center', horizontal='center')
                sheet1.cell(row, col).font = Font(size=10, bold=True, italic=True)
                sheet1.cell(row, col).border = Border(bottom=Side(style='thin', color='000000'))
                col += 1
                sheet1.cell(row, col, "Door Name")
                sheet1.cell(row, col).alignment = Alignment(wrapText=True, vertical='center', horizontal='center')
                sheet1.cell(row, col).font = Font(size=10, bold=True, italic=True)
                sheet1.cell(row, col).border = Border(bottom=Side(style='thin', color='000000'))
                col += 1
                sheet1.cell(row, col, "Comment")
                sheet1.cell(row, col).alignment = Alignment(wrapText=True, vertical='center', horizontal='center')
                sheet1.cell(row, col).font = Font(size=10, bold=True, italic=True)
                sheet1.cell(row, col).border = Border(bottom=Side(style='thin', color='000000'))
                col += 1
                sheet1.cell(row, col, "W")
                sheet1.cell(row, col).alignment = Alignment(wrapText=True, vertical='center', horizontal='center')
                sheet1.cell(row, col).font = Font(size=10, bold=True, italic=True)
                sheet1.cell(row, col).border = Border(bottom=Side(style='thin', color='000000'))
                col += 1
                sheet1.cell(row, col, "H")
                sheet1.cell(row, col).alignment = Alignment(wrapText=True, vertical='center', horizontal='center')
                sheet1.cell(row, col).font = Font(size=10, bold=True, italic=True)
                sheet1.cell(row, col).border = Border(bottom=Side(style='thin', color='000000'))
                col += 1
                sheet1.cell(row, col, "Std")
                sheet1.cell(row, col).alignment = Alignment(wrapText=True, vertical='center', horizontal='center')
                sheet1.cell(row, col).font = Font(size=10, bold=True, italic=True)
                sheet1.cell(row, col).border = Border(bottom=Side(style='thin', color='000000'))
                col += 1
                sheet1.cell(row, col, "Edge")
                sheet1.cell(row, col).alignment = Alignment(wrapText=True, vertical='center', horizontal='center')
                sheet1.cell(row, col).font = Font(size=10, bold=True, italic=True)
                sheet1.cell(row, col).border = Border(bottom=Side(style='thin', color='000000'))
                col += 1
                sheet1.cell(row, col, "B/L")
                sheet1.cell(row, col).alignment = Alignment(wrapText=True, vertical='center', horizontal='center')
                sheet1.cell(row, col).font = Font(size=10, bold=True, italic=True)
                sheet1.cell(row, col).border = Border(bottom=Side(style='thin', color='000000'))
                col += 1
                sheet1.cell(row, col, "B2/L2")
                sheet1.cell(row, col).alignment = Alignment(wrapText=True, vertical='center', horizontal='center')
                sheet1.cell(row, col).font = Font(size=10, bold=True, italic=True)
                sheet1.cell(row, col).border = Border(bottom=Side(style='thin', color='000000'))
                col += 1
                sheet1.cell(row, col, "B3/L3")
                sheet1.cell(row, col).alignment = Alignment(wrapText=True, vertical='center', horizontal='center')
                sheet1.cell(row, col).font = Font(size=10, bold=True, italic=True)
                sheet1.cell(row, col).border = Border(bottom=Side(style='thin', color='000000'))
                col += 1
                sheet1.cell(row, col, "T/R")
                sheet1.cell(row, col).alignment = Alignment(wrapText=True, vertical='center', horizontal='center')
                sheet1.cell(row, col).font = Font(size=10, bold=True, italic=True)
                sheet1.cell(row, col).border = Border(bottom=Side(style='thin', color='000000'))
                col += 1
                sheet1.cell(row, col, "Cab #")
                sheet1.cell(row, col).alignment = Alignment(wrapText=True, vertical='center', horizontal='center')
                sheet1.cell(row, col).font = Font(size=10, bold=True, italic=True)
                sheet1.cell(row, col).border = Border(bottom=Side(style='thin', color='000000'))
                row += 1
                col -= 11
                type_total = 0

                hinge.sort(key = lambda x: x[10])
                hinge.sort(key = lambda x: x[7])
                hinge.sort(key = lambda x: x[3])
                hinge.sort(key = lambda x: x[4])
                hinge.sort(key = lambda x: x[6])
                hinge.sort(key = lambda x: x[5])
                
                for door in hinge:
                    if door[7] == 0 and door[8] == 0 and door[9] == 0 and door[10] == 0:
                        continue
                    # Quantity
                    sheet1.cell(row, col, door[0])
                    sheet1.cell(row, col).alignment = Alignment(wrapText=True, vertical='center', horizontal='center')
                    sheet1.cell(row, col).font = Font(size=9)
                    sheet1.cell(row, col).border = Border(bottom=Side(style='thin', color='D4D4D4'))
                    col += 1
                    # Name
                    sheet1.cell(row, col, door[1])
                    sheet1.cell(row, col).alignment = Alignment(wrapText=True, vertical='center', horizontal='center')
                    sheet1.cell(row, col).font = Font(size=9)
                    sheet1.cell(row, col).border = Border(bottom=Side(style='thin', color='D4D4D4'))
                    col += 1
                    # Comment
                    sheet1.cell(row, col, door[2])
                    sheet1.cell(row, col).alignment = Alignment(wrapText=True, vertical='center', horizontal='center')
                    sheet1.cell(row, col).font = Font(size=9)
                    sheet1.cell(row, col).border = Border(bottom=Side(style='thin', color='D4D4D4'))
                    col += 1
                    # Width
                    sheet1.cell(row, col, round(door[3],1))
                    sheet1.cell(row, col).alignment = Alignment(wrapText=True, vertical='center', horizontal='center')
                    sheet1.cell(row, col).font = Font(size=9)
                    sheet1.cell(row, col).border = Border(bottom=Side(style='thin', color='D4D4D4'))
                    col += 1
                    # Height
                    sheet1.cell(row, col, round(door[4],1))
                    sheet1.cell(row, col).alignment = Alignment(wrapText=True, vertical='center', horizontal='center')
                    sheet1.cell(row, col).font = Font(size=9)
                    sheet1.cell(row, col).border = Border(bottom=Side(style='thin', color='D4D4D4'))
                    col += 1
                    # Standard or Non-Standard
                    sheet1.cell(row, col, door[5])
                    sheet1.cell(row, col).alignment = Alignment(wrapText=True, vertical='center', horizontal='center')
                    sheet1.cell(row, col).font = Font(size=9)
                    sheet1.cell(row, col).border = Border(bottom=Side(style='thin', color='D4D4D4'))
                    col += 1
                    # Hinge Edge
                    sheet1.cell(row, col, door[6])
                    sheet1.cell(row, col).alignment = Alignment(wrapText=True, vertical='center', horizontal='center')
                    sheet1.cell(row, col).font = Font(size=9)
                    sheet1.cell(row, col).border = Border(bottom=Side(style='thin', color='D4D4D4'))
                    col += 1
                    # Bottom/Left Hinge
                    hinge_center = door[7] if door[7] != 0 else ''
                    sheet1.cell(row, col, hinge_center)
                    sheet1.cell(row, col).alignment = Alignment(wrapText=True, vertical='center', horizontal='center')
                    sheet1.cell(row, col).font = Font(size=9)
                    sheet1.cell(row, col).border = Border(bottom=Side(style='thin', color='D4D4D4'))
                    col += 1
                    # Bottom/Left Mid Hinge
                    hinge_center = door[8] if door[8] != 0 else ''
                    sheet1.cell(row, col, hinge_center)
                    sheet1.cell(row, col).alignment = Alignment(wrapText=True, vertical='center', horizontal='center')
                    sheet1.cell(row, col).font = Font(size=9)
                    sheet1.cell(row, col).border = Border(bottom=Side(style='thin', color='D4D4D4'))
                    col += 1
                    # Top/Right Mid Hinge
                    hinge_center = door[9] if door[9] != 0 else ''
                    sheet1.cell(row, col, hinge_center)
                    sheet1.cell(row, col).alignment = Alignment(wrapText=True, vertical='center', horizontal='center')
                    sheet1.cell(row, col).font = Font(size=9)
                    sheet1.cell(row, col).border = Border(bottom=Side(style='thin', color='D4D4D4'))
                    col += 1
                    # Top/Right Hinge
                    hinge_center = door[10] if door[10] != 0 else ''
                    sheet1.cell(row, col, hinge_center)
                    sheet1.cell(row, col).alignment = Alignment(wrapText=True, vertical='center', horizontal='center')
                    sheet1.cell(row, col).font = Font(size=9)
                    sheet1.cell(row, col).border = Border(bottom=Side(style='thin', color='D4D4D4'))
                    col += 1
                    
                    # convert list of cab_nums to string
                    cab_num_string = ', '.join(door[11])
                    
                    # Cabinet Numbers
                    sheet1.cell(row, col, cab_num_string)
                    sheet1.cell(row, col).alignment = Alignment(wrapText=True, vertical='center', horizontal='center')
                    sheet1.cell(row, col).font = Font(size=9)
                    sheet1.cell(row, col).border = Border(bottom=Side(style='thin', color='D4D4D4'))
                    

                    # increment quantity
                    type_total += door[0]

                    # reset row & column
                    col -= 11
                    row += 1

                row -= 1
                sheet1.cell(row, col).border = Border(bottom=Side(style='thin', color='000000'))
                sheet1.cell(row, col + 1).border = Border(bottom=Side(style='thin', color='000000'))
                sheet1.cell(row, col + 2).border = Border(bottom=Side(style='thin', color='000000'))
                sheet1.cell(row, col + 3).border = Border(bottom=Side(style='thin', color='000000'))
                sheet1.cell(row, col + 4).border = Border(bottom=Side(style='thin', color='000000'))
                sheet1.cell(row, col + 5).border = Border(bottom=Side(style='thin', color='000000'))
                sheet1.cell(row, col + 6).border = Border(bottom=Side(style='thin', color='000000'))
                sheet1.cell(row, col + 7).border = Border(bottom=Side(style='thin', color='000000'))
                sheet1.cell(row, col + 8).border = Border(bottom=Side(style='thin', color='000000'))
                sheet1.cell(row, col + 9).border = Border(bottom=Side(style='thin', color='000000'))
                sheet1.cell(row, col + 10).border = Border(bottom=Side(style='thin', color='000000'))
                sheet1.cell(row, col + 11).border = Border(bottom=Side(style='thin', color='000000'))
                row += 1
                sheet1.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col + 11)
                sheet1.cell(row, col, "Hinge Type Total: (" + str(type_total) + ")")
                sheet1.cell(row, col).alignment = Alignment(wrapText=True, vertical='top', horizontal='left')
                sheet1.cell(row, col).font = Font(size=10, italic=True)

                style_total += type_total
                row += 2
        
            row += 1
            sheet1.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col + 11)
            sheet1.cell(row, col, "Door Style Total: (" + str(style_total) + ")")
            sheet1.cell(row, col).alignment = Alignment(wrapText=True, vertical='center', horizontal='left')
            sheet1.cell(row, col).font = Font(size=11, bold=True, italic=True)
            sheet1.cell(row, col).border = Border(bottom=Side(style='thin', color='000000'))
            sheet1.cell(row, col + 1).border = Border(bottom=Side(style='thin', color='000000'))
            sheet1.cell(row, col + 2).border = Border(bottom=Side(style='thin', color='000000'))
            col -= 1

            mat_total += style_total
            row += 2
            
        sheet1.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col + 11)
        sheet1.cell(row, col, "Material Total: (" + str(mat_total) + ")")
        sheet1.cell(row, col).alignment = Alignment(wrapText=True, vertical='center', horizontal='left')
        sheet1.cell(row, col).font = Font(size=14, italic=True, bold=True)
        page_break = Break(id=row)
        sheet1.row_breaks.append(page_break)
        row += 1
                
    
    sheet1.column_dimensions['A'].width = 4
    sheet1.column_dimensions['B'].width = 4
    sheet1.column_dimensions['C'].width = 11
    sheet1.column_dimensions['D'].width = 20
    sheet1.column_dimensions['E'].width = 7
    sheet1.column_dimensions['F'].width = 7
    sheet1.column_dimensions['G'].width = 4
    sheet1.column_dimensions['H'].width = 7
    sheet1.column_dimensions['I'].width = 7
    sheet1.column_dimensions['J'].width = 7
    sheet1.column_dimensions['K'].width = 7
    sheet1.column_dimensions['L'].width = 7
    sheet1.column_dimensions['M'].width = 20

    sheet1.page_margins.left = 0.5
    sheet1.page_margins.right = 0.5
    sheet1.page_margins.top = 0.5
    sheet1.page_margins.bottom = 0.5
    sheet1.page_margins.footer = 0.25


    sheet1.oddFooter.right.text = "Page &[Page] of &N"
    sheet1.oddFooter.right.size = 9
    sheet1.oddFooter.right.color = "000000"
    
    print_area = 'A1:M' + str(row - 1)
    sheet1.print_area = print_area
    sheet1.sheet_properties.pageSetUpPr.fitToPage = True
    sheet1.page_setup.fitToHeight = False   
    
    save_name = job_name + ' - Hinge Boring List' + '.xlsx'
    full_save_name = os.path.join(dir_path, save_name)
    try:
        wb.save(full_save_name)
    except PermissionError:
        print("\nSAVE FAILED\nYou will need to close the open file before it can be saved.")

    os.startfile(full_save_name)

    
hinge_boring_report()


input("Press Enter to Close")
