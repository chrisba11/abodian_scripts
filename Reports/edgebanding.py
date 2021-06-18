import os
from openpyxl.workbook import Workbook
from openpyxl.styles import Border, Side, Alignment, Font


# This is to ask for the directory path at the command prompt
dir_path = input('What is the full directory path? ')


def edgeband_report():
    """
    Generates an Excel file listing the banding templates used for a specified job.
    The function opens each of the .des files inside a job directory and looks 
        at every product in the room.
    If the product has a banding template that isn't in the dict,
        it adds that banding temp to a dictionary of templates.
    Then it writes data from the dictionary of templates to an Excel file
        inside the directory it was told to run in.
    It formats the cells as it goes and sets the print area the content inserted.
    Finally, it opens the Excel file so the user can print.
    """
    dir_path_lst = dir_path.split('\\')
    job_name = dir_path_lst[-1]

    temp_dict = {}
    
    for root, dirs, files in os.walk(dir_path):
        
        for file in files:

            if file.endswith('.des'):

                full_path = dir_path + '\\' + file
                f = open(full_path, "rt")
                content = f.readlines()
                f.close()
                room_num = file[4:file.find('.')]
                mat_temp_start_idx = content[4].find('MatBandingTemplate=') + 20
                mat_temp_end_idx = content[4].find('" CabBaseDoorTexture=')
                room_mat_temp = content[4][mat_temp_start_idx:mat_temp_end_idx]

                if room_mat_temp not in temp_dict:
                    temp_dict[room_mat_temp] = {"Banding":[], "Cabinets": []}
    
                for line in content:
                    if line.startswith('    <Product '):
                        mat_or_start_idx = line.find('MatBandingOR=') + 14
                        mat_or_end_idx = line.find('" LToeAdj=')
                        mat_or_intro = line[mat_or_start_idx:mat_or_start_idx + 2]

                        num_start_idx = line.find('CabNo=') + 7
                        num_end_idx = line.find(' Numbered=')
                        prod_num = line[num_start_idx:num_end_idx - 1]
                        is_numbered_start_idx = num_end_idx + 11
                        is_numbered = line[is_numbered_start_idx:is_numbered_start_idx + 1]
                        mat_temp_or = line[mat_or_start_idx:mat_or_end_idx]

                        if is_numbered == 'T':
                            full_prod_num = 'R' + room_num + 'C' + prod_num
                        else:
                            full_prod_num = 'R' + room_num + 'N' + prod_num

                        if mat_or_intro == '" ':
                            temp_dict[room_mat_temp]["Cabinets"].append(full_prod_num)
                        elif mat_temp_or in temp_dict:
                            temp_dict[mat_temp_or]["Cabinets"].append(full_prod_num)
                        else:
                            temp_dict[mat_temp_or] = {"Banding": [], "Cabinets": [full_prod_num]}

    if "None" in temp_dict:
        temp_dict.pop("None")

    data_path = '\\'.join(dir_path_lst[:-2])
    data_path += '\\Data'

    for template in temp_dict:

        for root, dirs, files in os.walk(data_path):
            
            for file in files:

                if file.startswith(template):

                    full_path = data_path + '\\' + file
                    f = open(full_path, "rt")
                    content = f.readlines()
                    f.close()

                    band_num = 1

                    lbl_sym_start_idx = content[2].find('SymbolForLabels=') + 17
                    label_symbol = '[ ' + content[2][lbl_sym_start_idx:lbl_sym_start_idx + 2] + ' ]'
                    temp_dict[template]['Symbol'] = label_symbol

                    for line in content[3:7]:
                        mat_start_idx = line.find('Mat=') + 5
                        mat_end_idx = line.find('" MatThick=')
                        mat_name = line[mat_start_idx:mat_end_idx]
                        temp_dict[template]["Banding"].append("EB" + str(band_num) +": " + mat_name)
                        band_num += 1


    temp_list = []

    for template in temp_dict:
        temp_list.append([temp_dict[template]['Symbol'], template])

    temp_list.sort()


    wb = Workbook()
    sheet1 = wb.active
    row = 1
    col = 1

    sheet1.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col + 1)
    sheet1.cell(row, col, job_name + ' - Edgebanding')
    sheet1.cell(row, col).alignment = Alignment(vertical='top')
    sheet1.cell(row, col).font = Font(size=12, bold=True, italic=True)
    sheet1.row_dimensions[1].height = 40
    row += 1

    for template in temp_list:
        sheet1.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col + 1)
        sheet1.cell(row, col, template[0] + ' ' + template[1])
        sheet1.cell(row, col).alignment = Alignment(wrapText=True, horizontal='left')
        sheet1.cell(row, col).font = Font(size=14, bold=True, underline='single')
        row += 1


        for banding in temp_dict[template[1]]["Banding"][:2]:
            sheet1.cell(row, col, banding)
            sheet1.cell(row, col).alignment = Alignment(wrapText=True, vertical='top', indent=1.0)
            sheet1.cell(row, col).font = Font(size=11)
            sheet1.cell(row, col).border = Border(bottom=Side(style='thin', color='D4D4D4'))
            row += 1

        col += 1
        row -= 2

        for banding in temp_dict[template[1]]["Banding"][2:]:
            sheet1.cell(row, col, banding)
            sheet1.cell(row, col).alignment = Alignment(wrapText=True, vertical='top', indent=1.0)
            sheet1.cell(row, col).font = Font(size=11)
            sheet1.cell(row, col).border = Border(bottom=Side(style='thin', color='D4D4D4'))
            row += 1

        col -= 1

        sheet1.cell(row, col, ", ".join(temp_dict[template[1]]["Cabinets"]))
        sheet1.cell(row, col).alignment = Alignment(wrapText=True, vertical="center", horizontal='left')
        sheet1.cell(row, col).font = Font(size=10, color="0000FF")
        sheet1.row_dimensions[row].height = 30 + (10 * (len(temp_dict[template[1]]["Cabinets"])//12))
        sheet1.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col + 1)

        row += 2

    sheet1.column_dimensions['A'].width = 43
    sheet1.column_dimensions['B'].width = 43
    
    print_area = 'A1:B' + str(row)
    sheet1.print_area = print_area
    
    save_name = job_name + ' - Edgebanding' + '.xlsx'
    full_save_name = os.path.join(dir_path, save_name)
    try:
        wb.save(full_save_name)
    except PermissionError:
        print("\nSAVE FAILED\nYou will need to close the open file before it can be saved.")

    os.startfile(full_save_name)
    
edgeband_report()

input("Press Enter to Close")